from flask import Flask, request, send_file, render_template
import pandas as pd
import io
from datetime import datetime

app = Flask(__name__)

ALLOWED_EXTENSIONS = {'.xlsx', '.xls'}

def is_excel(filename):
    return any(filename.lower().endswith(ext) for ext in ALLOWED_EXTENSIONS)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    warehouse_file = request.files.get('warehouse')
    branch_file = request.files.get('branch')

    if not warehouse_file or not branch_file:
        return 'يرجى رفع الملفين', 400

    if not is_excel(warehouse_file.filename):
        return 'ملف المستودع يجب أن يكون Excel (.xlsx أو .xls)', 400

    if not is_excel(branch_file.filename):
        return 'ملف الفرع يجب أن يكون Excel (.xlsx أو .xls)', 400

    try:
        warehouse = pd.read_excel(warehouse_file)
        branch = pd.read_excel(branch_file)
    except Exception:
        return 'تعذر قراءة أحد الملفات، تأكد أنهما ملفات Excel صحيحة', 400

    # تنظيف أسماء الأعمدة
    warehouse.columns = warehouse.columns.str.strip()
    branch.columns = branch.columns.str.strip()

    # 🔥 تحديد عمود الكود تلقائيًا
    def detect_code_column(df):
        possible = [col for col in df.columns if 'كود' in col or 'رقم' in col]
        for col in possible:
            if df[col].astype(str).str.replace('.', '').str.isnumeric().sum() > 5:
                return col
        return None

    warehouse_code_col = detect_code_column(warehouse)
    branch_code_col = detect_code_column(branch)

    if warehouse_code_col is None:
        return "❌ لم يتم تحديد عمود كود الصنف في ملف المستودع"

    if branch_code_col is None:
        return "❌ لم يتم تحديد عمود كود الصنف في ملف الفرع"

    # تحويل القيم لنص
    warehouse[warehouse_code_col] = warehouse[warehouse_code_col].astype(str).str.strip()
    branch[branch_code_col] = branch[branch_code_col].astype(str).str.strip()

    # 🔥 تحديد عمود الرصيد تلقائيًا
    def detect_qty_column(df):
        for col in df.columns:
            if 'رصيد' in col:
                return col
        return None

    qty_col = detect_qty_column(warehouse)

    if qty_col is None:
        return "❌ لم يتم العثور على عمود الرصيد في المستودع"

    # 🔥 تحديد عمود اسم الصنف (اختياري)
    def detect_name_column(df):
        for col in df.columns:
            if 'اسم' in col or 'صنف' in col:
                return col
        return None

    name_col = detect_name_column(warehouse)
    if name_col is None:
        name_col = warehouse.columns[1]  # fallback

    # الأصناف المتاحة في المستودع
    warehouse_available = warehouse[warehouse[qty_col] > 0]

    # أكواد الفرع
    branch_codes = set(branch[branch_code_col].unique())

    # الأصناف غير الموجودة في الفرع
    zero_at_branch = warehouse_available[
        ~warehouse_available[warehouse_code_col].isin(branch_codes)
    ]

    # بناء النتيجة
    result = zero_at_branch[[warehouse_code_col, name_col, qty_col]].copy()
    result.columns = ['كود الصنف', 'اسم الصنف', 'رصيد المستودع']
    result = result.sort_values('اسم الصنف').reset_index(drop=True)

    # كتابة Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result.to_excel(writer, index=False, sheet_name='الاحتياجات')

        workbook = writer.book
        worksheet = writer.sheets['الاحتياجات']

        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)

        for col in range(1, 4):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        worksheet.column_dimensions['A'].width = 18
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 18

        light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        for row in range(2, len(result) + 2):
            if row % 2 == 0:
                for col in range(1, 4):
                    worksheet.cell(row=row, column=col).fill = light_fill

    output.seek(0)
    today = datetime.now().strftime('%d-%m-%Y')
    filename = f'الاحتياجات_{today}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    app.run(debug=True, port=5050)
